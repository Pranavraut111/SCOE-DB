from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File, BackgroundTasks, Response
from fastapi.responses import Response
from sqlalchemy.orm import Session

from app.crud import crud_student
from app.schemas import student as student_schema
from app.services.generators import generate_roll_number, generate_institutional_email
from app.core.config import settings
import csv
import io
import smtplib
from email.mime.text import MIMEText
from sqlalchemy.exc import IntegrityError
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from app.db.session import get_db

router = APIRouter()

@router.post("/", response_model=student_schema.Student, status_code=status.HTTP_201_CREATED)
def create_student(
    *,
    db: Session = Depends(get_db),
    student_in: student_schema.StudentCreate,
):
    """
    Create new student.
    """
    # Check if email already exists
    student = crud_student.get_student_by_email(db, email=student_in.email)
    if student:
        raise HTTPException(
            status_code=400,
            detail="A student with this email already exists."
        )
    
    # Check if admission number already exists
    student = crud_student.get_student_by_admission_number(
        db, admission_number=student_in.admission_number
    )
    if student:
        raise HTTPException(
            status_code=400,
            detail="A student with this admission number already exists."
        )
    
    return crud_student.create_student(db=db, student=student_in)

@router.get("/import/template-new")
def download_new_template():
    """
    Download Excel template for bulk import with proper structure and sample data
    """
    import pandas as pd
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    # Create template with sample data from Computer Engineering students
    template_data = {
        'First Name': ['Aarav', 'Priya', 'Arjun', 'Sneha', 'Rohan'],
        'Middle Name': ['Rajesh', 'Suresh', 'Vikram', 'Anil', 'Prakash'],
        'Last Name': ['Sharma', 'Patel', 'Singh', 'Deshmukh', 'Kulkarni'],
        'Address': [
            '123 MG Road, Pune',
            '456 FC Road, Pune', 
            '789 JM Road, Pune',
            '321 Karve Road, Pune',
            '654 Baner Road, Pune'
        ],
        'Gender': ['Male', 'Female', 'Male', 'Female', 'Male'],
        'Category': ['General', 'OBC', 'General', 'SC', 'General'],
        'Date of Birth': ['2005-03-15', '2005-07-22', '2005-01-10', '2005-09-05', '2005-11-18'],
        'Phone Number': ['9876543210', '9876543211', '9876543212', '9876543213', '9876543214'],
        'Branch': [
            'Computer Science Engineering',
            'Computer Science Engineering', 
            'Computer Science Engineering',
            'Computer Science Engineering',
            'Computer Science Engineering'
        ],
        'Year': ['1st Year', '1st Year', '1st Year', '1st Year', '1st Year'],
        'Mother Name': ['Priya Sharma', 'Sunita Patel', 'Kavita Singh', 'Mangala Deshmukh', 'Shobha Kulkarni']
    }
    
    df = pd.DataFrame(template_data)
    
    # Create Excel file in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Students', index=False)
        
        # Add instructions sheet
        instructions = pd.DataFrame({
            'Column': ['First Name', 'Middle Name', 'Last Name', 'Address', 'Gender', 'Category', 'Date of Birth', 'Phone Number', 'Branch', 'Year', 'Mother Name'],
            'Required': ['Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes'],
            'Format/Options': [
                'Text',
                'Text', 
                'Text',
                'Full address',
                'Male or Female',
                'General, OBC, SC, or ST',
                'YYYY-MM-DD (e.g., 2000-01-15)',
                '10 digits',
                'Computer Science Engineering, Information Technology, Electronics and Communication Engineering, Electrical Engineering, Mechanical Engineering, Civil Engineering',
                '1st Year, 2nd Year, 3rd Year, 4th Year',
                'Full name of mother'
            ]
        })
        instructions.to_excel(writer, sheet_name='Instructions', index=False)
    
    output.seek(0)
    
    return StreamingResponse(
        BytesIO(output.read()),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": "attachment; filename=student_bulk_import_template_with_demo_data.xlsx", "Cache-Control": "no-cache, no-store, must-revalidate", "Pragma": "no-cache", "Expires": "0"}
    )

@router.post("/import/preview")
async def import_preview(file: UploadFile = File(...)):
    """
    Preview import file and validate structure
    """
    if not (file.filename.endswith('.csv') or file.filename.endswith('.xlsx')):
        raise HTTPException(status_code=400, detail="Only CSV and Excel (.xlsx) files are supported")
    
    try:
        content = await file.read()
        
        if file.filename.endswith('.xlsx'):
            try:
                # Try reading from different starting rows to skip metadata
                df = None
                for skip_rows in range(0, 10):  # Try up to 10 rows
                    try:
                        temp_df = pd.read_excel(BytesIO(content), skiprows=skip_rows)
                        # Check if this row contains our expected columns
                        if not temp_df.empty:
                            columns = temp_df.columns.tolist()
                            # Look for expected column patterns
                            has_first_name = any('First Name' in str(col) for col in columns)
                            has_name = any('Name' in str(col) and 'First' not in str(col) and 'Last' not in str(col) for col in columns)
                            has_address = any('Address' in str(col) for col in columns)
                            
                            if has_first_name or (has_name and has_address):
                                df = temp_df
                                break
                    except:
                        continue
                
                if df is None:
                    # Fallback to default reading
                    df = pd.read_excel(BytesIO(content))
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Error reading Excel file: {str(e)}")
        else:
            # Parse CSV file
            df = pd.read_csv(BytesIO(content))
        
        # Convert NaN to empty strings for consistent processing
        df = df.fillna('')
        rows = df.to_dict('records')
        
        # Required columns in exact order
        required_columns = [
            'First Name', 'Middle Name', 'Last Name', 'Address', 'Gender', 
            'Category', 'Date of Birth', 'Phone Number', 'Branch', 'Year', 'Mother Name'
        ]
        
        # Check if file has any data
        if not rows:
            raise HTTPException(status_code=400, detail="File is empty or has no data rows")
        
        # Validate column headers
        file_columns = list(rows[0].keys()) if rows else []
        missing_columns = [col for col in required_columns if col not in file_columns]
        
        if missing_columns:
            raise HTTPException(
                status_code=400, 
                detail=f"Missing required columns: {', '.join(missing_columns)}. Please use the provided template."
            )
        
        # Validate each row
        preview = []
        valid_genders = ['male', 'female']
        valid_categories = ['general', 'obc', 'sc', 'st']
        valid_branches = [
            'Computer Science Engineering', 'Information Technology',
            'Electronics and Communication Engineering', 'Electrical Engineering',
            'Mechanical Engineering', 'Civil Engineering'
        ]
        valid_years = ['1st Year', '2nd Year', '3rd Year', '4th Year']
        
        for idx, row in enumerate(rows, 1):
            errors = []
            
            # Check required fields
            for col in required_columns:
                value = str(row.get(col, '')).strip()
                if not value:
                    errors.append(f"Row {idx}: Missing {col}")
            
            # Validate specific fields
            gender = str(row.get('Gender', '')).strip().lower()
            if gender and gender not in valid_genders:
                errors.append(f"Row {idx}: Gender must be 'Male' or 'Female'")
            
            category = str(row.get('Category', '')).strip().lower()
            if category and category not in valid_categories:
                errors.append(f"Row {idx}: Category must be one of: General, OBC, SC, ST")
            
            branch = str(row.get('Branch', '')).strip()
            if branch and branch not in valid_branches:
                errors.append(f"Row {idx}: Invalid branch. Use template values.")
            
            year = str(row.get('Year', '')).strip()
            if year and year not in valid_years:
                errors.append(f"Row {idx}: Year must be one of: 1st Year, 2nd Year, 3rd Year, 4th Year")
            
            # Validate date format
            dob = str(row.get('Date of Birth', '')).strip()
            if dob:
                try:
                    from datetime import datetime
                    datetime.strptime(dob, '%Y-%m-%d')
                except ValueError:
                    errors.append(f"Row {idx}: Date of Birth must be in YYYY-MM-DD format")
            
            # Validate phone number
            phone = str(row.get('Phone Number', '')).strip()
            if phone and (not phone.isdigit() or len(phone) != 10):
                errors.append(f"Row {idx}: Phone Number must be 10 digits")
            
            preview.append({
                'row_number': idx,
                'data': row,
                'is_valid': len(errors) == 0,
                'errors': errors
            })
        
        valid_count = sum(1 for p in preview if p['is_valid'])
        
        return {
            'total_rows': len(rows),
            'valid_rows': valid_count,
            'invalid_rows': len(rows) - valid_count,
            'preview': preview[:10],  # Show first 10 rows for preview
            'can_import': valid_count > 0
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")

@router.post("/import/save")
async def import_save(
    *,
    db: Session = Depends(get_db),
    file: UploadFile = File(...),
):
    """
    Process and save bulk import data
    """
    if not (file.filename.endswith('.csv') or file.filename.endswith('.xlsx')):
        raise HTTPException(status_code=400, detail="Only CSV and Excel (.xlsx) files are supported")
    
    try:
        content = await file.read()
        
        if file.filename.endswith('.xlsx'):
            # Parse Excel file - handle Numbers export metadata
            try:
                # Try reading from different starting rows to skip metadata
                df = None
                for skip_rows in range(0, 10):  # Try up to 10 rows
                    try:
                        temp_df = pd.read_excel(io.BytesIO(content), skiprows=skip_rows)
                        # Check if this row contains our expected columns
                        if not temp_df.empty:
                            columns = temp_df.columns.tolist()
                            # Look for expected column patterns
                            has_first_name = any('First Name' in str(col) for col in columns)
                            has_name = any('Name' in str(col) and 'First' not in str(col) and 'Last' not in str(col) for col in columns)
                            has_address = any('Address' in str(col) for col in columns)
                            
                            if has_first_name or (has_name and has_address):
                                df = temp_df
                                break
                    except:
                        continue
                
                if df is None:
                    # Fallback to default reading
                    df = pd.read_excel(io.BytesIO(content))
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Error reading Excel file: {str(e)}")
            
            df = df.fillna('')  # Replace NaN with empty strings
            rows = df.to_dict('records')
        else:
            text = content.decode('utf-8')
            reader = csv.DictReader(io.StringIO(text))
            rows = list(reader)
        
        if not rows:
            raise HTTPException(status_code=400, detail="No data found in file")
        
        # Department to range mapping for sequential roll numbers
        dept_ranges = {
            'Computer Science Engineering': 1000,
            'Information Technology': 3000,
            'Electronics and Communication Engineering': 5000,
            'Electrical Engineering': 7000,
            'Mechanical Engineering': 8000,
            'Civil Engineering': 9000
        }
        
        total = 0
        successful = 0
        failed = 0
        errors = []
        created_students = []
        
        for idx, row in enumerate(rows, 1):
            total += 1
            try:
                # Extract and validate required fields
                first_name = str(row.get('First Name', '')).strip()
                middle_name = str(row.get('Middle Name', '')).strip()
                last_name = str(row.get('Last Name', '')).strip()
                address = str(row.get('Address', '')).strip()
                gender = str(row.get('Gender', '')).strip().lower()
                category = str(row.get('Category', '')).strip()
                dob = str(row.get('Date of Birth', '')).strip()
                phone = str(row.get('Phone Number', '')).strip()
                branch = str(row.get('Branch', '')).strip()
                year = str(row.get('Year', '')).strip()
                mother_name = str(row.get('Mother Name', '')).strip()
                
                # Validate required fields
                if not all([first_name, middle_name, last_name, address, gender, category, dob, phone, branch, year, mother_name]):
                    errors.append(f"Row {idx}: Missing required fields")
                    failed += 1
                    continue
                
                # Generate sequential roll number
                start_range = dept_ranges.get(branch, 1000)
                
                # Get existing students in this department
                existing_students = crud_student.get_students(db)
                dept_students = [s for s in existing_students if s.department == branch]
                
                # Find highest roll number in department range
                max_number = start_range
                for student in dept_students:
                    if student.roll_number and student.roll_number.startswith('SCOE'):
                        try:
                            num = int(student.roll_number[4:])
                            if start_range <= num < start_range + 1000 and num > max_number:
                                max_number = num
                        except ValueError:
                            continue
                
                next_roll_number = f"SCOE{max_number + 1:04d}"
                
                # Generate emails
                full_name = f"{first_name} {middle_name} {last_name}"
                personal_email = f"{first_name.lower()}.{last_name.lower()}@gmail.com"
                institutional_email = generate_institutional_email(full_name, branch)
                
                # Ensure unique emails and phone numbers
                email_suffix = 1
                original_personal = personal_email
                while crud_student.get_student_by_email(db, personal_email):
                    name_part = original_personal.split('@')[0]
                    personal_email = f"{name_part}{email_suffix}@gmail.com"
                    email_suffix += 1
                
                # Ensure unique institutional emails
                original_institutional = institutional_email
                inst_suffix = 1
                existing_students = crud_student.get_students(db)
                while any(s.institutional_email == institutional_email for s in existing_students):
                    # Add number to institutional email
                    email_parts = original_institutional.split('@')
                    institutional_email = f"{email_parts[0]}{inst_suffix}@{email_parts[1]}"
                    inst_suffix += 1
                    if inst_suffix > 100:  # Prevent infinite loop
                        break
                
                # Check for duplicate phone numbers and modify if needed
                original_phone = phone
                phone_suffix = 1
                while any(s.phone == phone for s in existing_students):
                    # Modify last digit to make unique
                    if len(original_phone) == 10 and original_phone.isdigit():
                        last_digit = int(original_phone[-1])
                        new_last_digit = (last_digit + phone_suffix) % 10
                        phone = original_phone[:-1] + str(new_last_digit)
                        phone_suffix += 1
                    else:
                        phone = f"{original_phone}{phone_suffix}"
                        phone_suffix += 1
                    if phone_suffix > 10:  # Prevent infinite loop
                        break
                
                # Validate and convert gender
                if gender not in ['male', 'female']:
                    errors.append(f"Row {idx}: Invalid gender '{gender}'")
                    failed += 1
                    continue
                
                # Create student record
                student_data = student_schema.StudentCreate(
                    first_name=first_name,
                    middle_name=middle_name,
                    last_name=last_name,
                    email=personal_email,
                    phone=phone,
                    date_of_birth=dob,
                    gender=student_schema.Gender[gender],
                    address=address,
                    state=year,
                    country='India',
                    admission_number=next_roll_number,
                    roll_number=next_roll_number,
                    institutional_email=institutional_email,
                    department=branch,
                    category=category,
                    mother_name=mother_name
                )
                
                # Save to database
                new_student = crud_student.create_student(db, student_data)
                created_students.append({
                    'name': full_name,
                    'roll_number': next_roll_number,
                    'department': branch
                })
                successful += 1
                
            except Exception as e:
                failed += 1
                errors.append(f"Row {idx}: {str(e)}")
        
        return {
            'total': total,
            'successful': successful,
            'failed': failed,
            'errors': errors[:10],  # Limit errors shown
            'created_students': created_students[:5]  # Show first 5 created students
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")

@router.get("/export/csv")
def export_students_csv(
    db: Session = Depends(get_db),
):
    """
    Export all students to CSV with proper structure
    """
    students = crud_student.get_students(db, limit=10000)
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write headers matching import template
    headers = [
        'First Name', 'Middle Name', 'Last Name', 'Address', 'Gender', 
        'Category', 'Date of Birth', 'Phone Number', 'Branch', 'Year', 
        'Mother Name', 'Roll Number', 'Personal Email', 'Institutional Email', 'Created At'
    ]
    writer.writerow(headers)
    
    # Write student data
    for student in students:
        writer.writerow([
            student.first_name or '',
            student.middle_name or '',
            student.last_name or '',
            student.address or '',
            student.gender.capitalize() if student.gender else '',
            student.category or 'General',
            student.date_of_birth.strftime('%Y-%m-%d') if student.date_of_birth else '',
            student.phone or '',
            student.department or '',
            student.state or '',
            student.mother_name or '',
            student.roll_number or '',
            student.email or '',
            student.institutional_email or '',
            student.created_at.strftime('%Y-%m-%d %H:%M:%S') if student.created_at else ''
        ])
    
    output.seek(0)
    
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8')),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=students_export.csv"}
    )

@router.get("/export/excel")
def export_students_excel(db: Session = Depends(get_db)):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        students = crud_student.get_students(db)
        
        # Create a new workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        
        # Define headers
        headers = [
            'Roll Number', 'First Name', 'Middle Name', 'Last Name', 'Institutional Email', 
            'Personal Email', 'Phone', 'Address', 'Gender', 'Department', 'Year', 'Mother Name'
        ]
        
        # Add headers to first row
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Add student data
        for row, student in enumerate(students, 2):
            ws.cell(row=row, column=1, value=getattr(student, 'roll_number', '') or '')
            ws.cell(row=row, column=2, value=student.first_name or '')
            ws.cell(row=row, column=3, value=getattr(student, 'middle_name', '') or '')
            ws.cell(row=row, column=4, value=student.last_name or '')
            ws.cell(row=row, column=5, value=getattr(student, 'institutional_email', '') or '')
            ws.cell(row=row, column=6, value=student.email or '')
            ws.cell(row=row, column=7, value=student.phone or '')
            ws.cell(row=row, column=8, value=student.address or '')
            ws.cell(row=row, column=9, value=student.gender.value if student.gender else '')
            ws.cell(row=row, column=10, value=getattr(student, 'department', '') or '')
            ws.cell(row=row, column=11, value=getattr(student, 'state', '') or '')
            ws.cell(row=row, column=12, value=getattr(student, 'mother_name', '') or '')
        
        # Style the header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Set column widths
        column_widths = [15, 15, 15, 15, 35, 30, 15, 40, 10, 35, 12, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=students.xlsx"}
        )
        
    except Exception as e:
        print(f"Excel export error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to export Excel file: {str(e)}")

@router.post("/email/csv")
def email_students_csv(
    *,
    db: Session = Depends(get_db),
    background: BackgroundTasks,
    recipient: str,
):
    def _send(csv_content: str):
        if not settings.SMTP_USER or not settings.SMTP_PASSWORD or not settings.SMTP_FROM_EMAIL:
            return
        msg = MIMEText(csv_content)
        msg['Subject'] = 'Students Export'
        msg['From'] = settings.SMTP_FROM_EMAIL
        msg['To'] = recipient
        with smtplib.SMTP(settings.SMTP_HOST, settings.SMTP_PORT) as server:
            server.starttls()
            server.login(settings.SMTP_USER, settings.SMTP_PASSWORD)
            server.sendmail(settings.SMTP_FROM_EMAIL, [recipient], msg.as_string())

    csv_content = export_students_csv(db)
    background.add_task(_send, csv_content)
    return { 'status': 'scheduled' }

@router.get("/", response_model=List[student_schema.Student])
def read_students(
    db: Session = Depends(get_db),
    skip: int = 0,
    limit: int = 100,
    search: Optional[str] = Query(None, description="Search by name, email, or admission number"),
):
    """
    Retrieve students with optional search and pagination.
    """
    students = crud_student.get_students(
        db, skip=skip, limit=limit, search=search
    )
    return students

@router.get("/count", response_model=int)
def get_students_count(
    db: Session = Depends(get_db),
):
    """
    Get total count of students.
    """
    return crud_student.get_students_count(db)

@router.get("/next-roll-number")
def get_next_roll_number(
    *,
    db: Session = Depends(get_db),
    department: str = Query(..., description="Department name"),
    start_range: int = Query(..., description="Starting range for department"),
):
    """
    Get the next sequential roll number for a department.
    """
    # Get all students in this department
    students = crud_student.get_students(db)
    department_students = [s for s in students if s.department == department]
    
    # Find the highest roll number in the department's range
    max_number = start_range
    for student in department_students:
        if student.roll_number and student.roll_number.startswith('SCOE'):
            try:
                num = int(student.roll_number[4:])  # Remove 'SCOE' prefix
                if start_range <= num < start_range + 1000 and num > max_number:
                    max_number = num
            except ValueError:
                continue
    
    next_number = max_number + 1
    return {"roll_number": f"{next_number:04d}"}

@router.get("/")
def read_students(
    skip: int = 0,
    limit: int = 100,
    search: str = None,
    db: Session = Depends(get_db)
):
    """
    Retrieve students.
    """
    students = crud_student.get_students(db, skip=skip, limit=limit, search=search)
    return students

@router.get("/{student_id}", response_model=student_schema.Student)
def read_student(
    student_id: int,
    db: Session = Depends(get_db),
):
    """
    Get student by ID.
    """
    student = crud_student.get_student(db, student_id=student_id)
    if not student:
        raise HTTPException(
            status_code=404,
            detail="Student not found"
        )
    return student

@router.put("/{student_id}", response_model=student_schema.Student)
def update_student(
    *,
    db: Session = Depends(get_db),
    student_id: int,
    student_in: student_schema.StudentUpdate,
):
    """
    Update a student.
    """
    student = crud_student.get_student(db, student_id=student_id)
    if not student:
        raise HTTPException(
            status_code=404,
            detail="Student not found"
        )
    
    # Check if email is being updated to an existing email
    if student_in.email and student_in.email != student.email:
        existing_student = crud_student.get_student_by_email(
            db, email=student_in.email
        )
        if existing_student:
            raise HTTPException(
                status_code=400,
                detail="A student with this email already exists."
            )
    
    # Check if admission number is being updated to an existing one
    if student_in.admission_number and student_in.admission_number != student.admission_number:
        existing_student = crud_student.get_student_by_admission_number(
            db, admission_number=student_in.admission_number
        )
        if existing_student:
            raise HTTPException(
                status_code=400,
                detail="A student with this admission number already exists."
            )
    
    return crud_student.update_student(
        db, db_student=student, student_in=student_in
    )

@router.delete("/{student_id}")
def delete_student(
    *,
    db: Session = Depends(get_db),
    student_id: int,
):
    """
    Delete a student.
    """
    student = crud_student.get_student(db, student_id=student_id)
    if not student:
        raise HTTPException(
            status_code=404,
            detail="Student not found"
        )
    return crud_student.delete_student(db=db, student_id=student_id)

@router.get("/demo/template-data")
def get_demo_students_for_template(
    db: Session = Depends(get_db),
    department: str = Query("Computer Engineering", description="Department to filter demo students"),
    limit: int = Query(5, description="Number of demo students to return")
):
    """
    Get demo students data for template generation
    """
    from app.models.student import Student
    
    # Get demo students from database
    demo_students = db.query(Student).filter(
        Student.department == department,
        Student.roll_number.like("SCOE%")
    ).limit(limit).all()
    
    if not demo_students:
        # Fallback to hardcoded data if no demo students in DB
        return [
            {
                "roll_number": "SCOE101",
                "first_name": "Aarav",
                "middle_name": "Rajesh", 
                "last_name": "Sharma",
                "email": "aarav.sharma@gmail.com",
                "phone": "9876543210",
                "department": department,
                "year": "1st Year"
            },
            {
                "roll_number": "SCOE102",
                "first_name": "Priya",
                "middle_name": "Suresh",
                "last_name": "Patel", 
                "email": "priya.patel@gmail.com",
                "phone": "9876543211",
                "department": department,
                "year": "1st Year"
            }
        ]
    
    # Convert to template format
    template_data = []
    for student in demo_students:
        template_data.append({
            "roll_number": student.roll_number,
            "first_name": student.first_name,
            "middle_name": student.middle_name,
            "last_name": student.last_name,
            "email": student.email,
            "phone": student.phone,
            "department": student.department,
            "year": "1st Year"  # Default for demo students
        })
    
    return template_data
